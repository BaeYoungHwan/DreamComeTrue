import io

from openpyxl import load_workbook

from src.core.db import get_connection, init_db
from src.export.report import (
    all_channels_settlement_to_excel_bytes,
    detailed_period_report,
    detailed_report_to_excel_bytes,
    export_all_data_to_excel_bytes,
    period_report,
    report_to_excel_bytes,
    settlement_to_excel_bytes,
)
from src.inventory.items import create_item
from src.inventory.variants import create_variant
from src.core.migrations import run_migrations
from src.settlement.channels import create_channel


def _seed_sale(conn, item_id, buyer, quantity, unit_price, sold_at, channel_id=None, variant_id=None):
    total = quantity * unit_price
    conn.execute(
        """
        INSERT INTO sales (item_id, channel_id, variant_id, buyer, quantity, unit_price, total_amount, sold_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, channel_id, variant_id, buyer, quantity, unit_price, total, sold_at),
    )
    conn.commit()


def test_period_report_aggregates_within_date_range(tmp_path):
    conn = get_connection(str(tmp_path / "farm.db"))
    init_db(conn)
    item_id = create_item(conn, "상추", "kg")

    _seed_sale(conn, item_id, "로컬푸드 매장", 3, 2000, "2026-07-10 09:00:00")
    _seed_sale(conn, item_id, "직거래", 2, 2500, "2026-07-15 09:00:00")
    _seed_sale(conn, item_id, "직거래", 1, 2500, "2026-08-01 09:00:00")

    report = period_report(conn, "2026-07-01", "2026-07-31")
    conn.close()

    assert report == [
        {"item_name": "상추", "total_quantity": 5, "total_amount": 11000}
    ]


def test_detailed_period_report_returns_individual_sale_rows_within_range(tmp_path):
    conn = get_connection(str(tmp_path / "farm.db"))
    init_db(conn)
    item_id = create_item(conn, "블루베리", "500g")
    variant_id = create_variant(conn, item_id, "특", "500g", 15000)
    channel_id = create_channel(conn, "모현점", "consignment", 10)

    _seed_sale(
        conn, item_id, "모현점", 3, 15000, "2026-07-10 09:00:00",
        channel_id=channel_id, variant_id=variant_id,
    )
    _seed_sale(conn, item_id, "직거래", 2, 2500, "2026-07-15 09:00:00")
    _seed_sale(conn, item_id, "직거래", 1, 2500, "2026-08-01 09:00:00")  # 기간 밖

    rows = detailed_period_report(conn, "2026-07-01", "2026-07-31")
    conn.close()

    assert rows == [
        {
            "sold_at": "2026-07-10 09:00:00",
            "item_name": "블루베리",
            "size": "특",
            "weight": "500g",
            "buyer": "모현점",
            "quantity": 3,
            "unit_price": 15000,
            "total_amount": 45000,
        },
        {
            "sold_at": "2026-07-15 09:00:00",
            "item_name": "블루베리",
            "size": None,
            "weight": None,
            "buyer": "직거래",
            "quantity": 2,
            "unit_price": 2500,
            "total_amount": 5000,
        },
    ]


def test_detailed_report_to_excel_bytes_contains_expected_rows():
    rows = [
        {
            "sold_at": "2026-07-10 09:00:00",
            "item_name": "블루베리",
            "size": "특",
            "weight": "500g",
            "buyer": "모현점",
            "quantity": 3,
            "unit_price": 15000,
            "total_amount": 45000,
        }
    ]

    data = detailed_report_to_excel_bytes(rows)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    assert [cell.value for cell in ws[1]] == [
        "날짜", "품목", "크기", "무게", "출하처", "수량", "단가", "금액"
    ]
    assert [cell.value for cell in ws[2]] == [
        "2026-07-10 09:00:00", "블루베리", "특", "500g", "모현점", 3, 15000, 45000
    ]


def test_report_to_excel_bytes_contains_expected_rows():
    rows = [{"item_name": "상추", "total_quantity": 5, "total_amount": 11000}]

    data = report_to_excel_bytes(rows)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    assert [cell.value for cell in ws[1]] == ["품목", "출하량", "판매대금"]
    assert [cell.value for cell in ws[2]] == ["상추", 5, 11000]


def test_export_all_data_to_excel_bytes_creates_four_sheets(tmp_path):
    conn = get_connection(str(tmp_path / "farm.db"))
    init_db(conn)
    run_migrations(conn)
    create_item(conn, "상추", "kg")

    data = export_all_data_to_excel_bytes(conn)
    conn.close()
    wb = load_workbook(io.BytesIO(data))

    assert set(wb.sheetnames) == {"품목", "입출고내역", "판매내역", "채널"}


def test_export_all_data_includes_channel_sheet_rows(tmp_path):
    conn = get_connection(str(tmp_path / "farm.db"))
    init_db(conn)
    run_migrations(conn)
    create_channel(conn, "모현점", "consignment", 10)

    data = export_all_data_to_excel_bytes(conn)
    conn.close()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["채널"]

    assert [cell.value for cell in ws[1]] == ["id", "채널명", "유형", "수수료율", "등록일"]
    assert ws[2][1].value == "모현점"


def test_settlement_to_excel_bytes_contains_expected_row():
    result = {
        "channel_name": "모현점",
        "period_start": "2026-06-01",
        "period_end": "2026-06-30",
        "sales_total": 293400,
        "commission_rate": 10,
        "commission_amount": 29340,
        "expected_deposit": 264060,
        "actual_deposit": 260800,
        "diff": -3260,
    }

    data = settlement_to_excel_bytes(result)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    assert ws["A1"].value == "채널"
    header = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]
    assert header == ["채널", "기간", "판매누계", "수수료율(%)", "수수료", "예상입금액", "실입금액", "차액"]
    assert row == [
        "모현점",
        "2026-06-01 ~ 2026-06-30",
        293400,
        10,
        29340,
        264060,
        260800,
        -3260,
    ]


def test_all_channels_settlement_to_excel_bytes_contains_one_row_per_channel():
    results = [
        {
            "channel_name": "모현점",
            "period_start": "2026-06-01",
            "period_end": "2026-06-30",
            "sales_total": 293400,
            "commission_rate": 10,
            "commission_amount": 29340,
            "expected_deposit": 264060,
            "actual_deposit": None,
            "diff": None,
        },
        {
            "channel_name": "어양점",
            "period_start": "2026-06-01",
            "period_end": "2026-06-30",
            "sales_total": 120000,
            "commission_rate": 10,
            "commission_amount": 12000,
            "expected_deposit": 108000,
            "actual_deposit": None,
            "diff": None,
        },
    ]

    data = all_channels_settlement_to_excel_bytes(results)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    assert header == ["채널", "기간", "판매누계", "수수료율(%)", "수수료", "예상입금액", "실입금액", "차액"]
    assert ws[2][0].value == "모현점"
    assert ws[3][0].value == "어양점"
    assert ws[3][2].value == 120000
