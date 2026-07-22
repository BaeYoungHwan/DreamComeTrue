"""기간별 정산 리포트 집계 및 엑셀(.xlsx) Export."""
import io
import sqlite3
from typing import Any

from openpyxl import Workbook


def _as_int(value: Any) -> Any:
    """금액 값을 정수로 변환한다(None은 그대로 반환)."""
    return None if value is None else round(value)


def period_report(
    conn: sqlite3.Connection, start_date: str, end_date: str
) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT items.name,
               SUM(sales.quantity) AS total_quantity,
               SUM(sales.total_amount) AS total_amount
        FROM sales
        JOIN items ON items.id = sales.item_id
        WHERE date(sales.sold_at) BETWEEN date(?) AND date(?)
        GROUP BY items.name
        ORDER BY items.name
        """,
        (start_date, end_date),
    ).fetchall()
    return [
        {"item_name": r[0], "total_quantity": r[1], "total_amount": r[2]}
        for r in rows
    ]


def detailed_period_report(
    conn: sqlite3.Connection, start_date: str, end_date: str
) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT sales.sold_at, items.name, item_variants.size, item_variants.weight,
               sales.buyer, sales.quantity, sales.unit_price, sales.total_amount
        FROM sales
        JOIN items ON items.id = sales.item_id
        LEFT JOIN item_variants ON item_variants.id = sales.variant_id
        WHERE date(sales.sold_at) BETWEEN date(?) AND date(?)
        ORDER BY sales.sold_at, sales.id
        """,
        (start_date, end_date),
    ).fetchall()
    return [
        {
            "sold_at": r[0],
            "item_name": r[1],
            "size": r[2],
            "weight": r[3],
            "buyer": r[4],
            "quantity": r[5],
            "unit_price": r[6],
            "total_amount": r[7],
        }
        for r in rows
    ]


def report_to_excel_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "정산리포트"
    ws.append(["품목", "출하량", "판매대금"])
    for row in rows:
        ws.append([row["item_name"], row["total_quantity"], _as_int(row["total_amount"])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detailed_report_to_excel_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "상세리포트"
    ws.append(["날짜", "품목", "크기", "무게", "출하처", "수량", "단가", "금액"])
    for row in rows:
        ws.append(
            [
                row["sold_at"],
                row["item_name"],
                row["size"],
                row["weight"],
                row["buyer"],
                row["quantity"],
                _as_int(row["unit_price"]),
                _as_int(row["total_amount"]),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def settlement_to_excel_bytes(result: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "채널정산"
    ws.append(
        ["채널", "기간", "판매누계", "수수료율(%)", "수수료", "예상입금액", "실입금액", "차액"]
    )
    ws.append(
        [
            result.get("channel_name"),
            f"{result.get('period_start')} ~ {result.get('period_end')}",
            _as_int(result.get("sales_total")),
            result.get("commission_rate"),
            _as_int(result.get("commission_amount")),
            _as_int(result.get("expected_deposit")),
            _as_int(result.get("actual_deposit")),
            _as_int(result.get("diff")),
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def all_channels_settlement_to_excel_bytes(results: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "전체채널정산"
    ws.append(
        ["채널", "기간", "판매누계", "수수료율(%)", "수수료", "예상입금액", "실입금액", "차액"]
    )
    for result in results:
        ws.append(
            [
                result.get("channel_name"),
                f"{result.get('period_start')} ~ {result.get('period_end')}",
                _as_int(result.get("sales_total")),
                result.get("commission_rate"),
                _as_int(result.get("commission_amount")),
                _as_int(result.get("expected_deposit")),
                _as_int(result.get("actual_deposit")),
                _as_int(result.get("diff")),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_all_data_to_excel_bytes(conn: sqlite3.Connection) -> bytes:
    wb = Workbook()

    ws_items = wb.active
    ws_items.title = "품목"
    ws_items.append(["id", "이름", "규격", "커스텀속성", "등록일"])
    for row in conn.execute(
        "SELECT id, name, unit, custom_attributes, created_at FROM items"
    ):
        ws_items.append(list(row))

    ws_tx = wb.create_sheet("입출고내역")
    ws_tx.append(["id", "품목ID", "거래유형", "수량", "일시"])
    for row in conn.execute(
        "SELECT id, item_id, type, quantity, created_at FROM stock_transactions"
    ):
        ws_tx.append(list(row))

    ws_sales = wb.create_sheet("판매내역")
    ws_sales.append(
        ["id", "품목ID", "채널ID", "출하처", "수량", "단가", "판매대금", "판매일시"]
    )
    for row in conn.execute(
        "SELECT id, item_id, channel_id, buyer, quantity, unit_price, total_amount, sold_at FROM sales"
    ):
        r = list(row)
        r[5] = _as_int(r[5])
        r[6] = _as_int(r[6])
        ws_sales.append(r)

    ws_channels = wb.create_sheet("채널")
    ws_channels.append(["id", "채널명", "유형", "수수료율", "등록일"])
    for row in conn.execute(
        "SELECT id, name, channel_type, commission_rate, created_at FROM channels"
    ):
        ws_channels.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
